import openpyxl

excel_path = r"C:\Users\User\Documents\Innovant\Internship\Project Dev\Development\AI Resume Analyser\testing report\InnovEdge_Plan_AI_Resume_Analyzer_RAG Report.xlsx"
wb = openpyxl.load_workbook(excel_path)
print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\nSheet: {sheetname}")
    # Print the first row (headers)
    first_row = [cell.value for cell in sheet[1]]
    print("Headers:", first_row)


