import openpyxl
import json

file_path = r'C:\Users\User\Documents\Innovant\Internship\Project Dev\Development\AI Resume Analyser\testing report\AI resume analyzer RAG report 20260708_183926.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

issues = []

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Find the header row (assume first row with data or first few rows)
    headers = {}
    header_row_idx = None
    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if any(cell and 'RAG' in str(cell) for cell in row):
            headers = {c_idx: cell for c_idx, cell in enumerate(row) if cell}
            header_row_idx = r_idx
            break
            
    if not header_row_idx:
        continue
        
    rag_col_idx = None
    desc_col_idx = None
    impact_col_idx = None
    reason_col_idx = None
    solution_col_idx = None
    
    for c_idx, header in headers.items():
        h = str(header).lower().strip()
        if 'rag' in h:
            rag_col_idx = c_idx
        elif 'acceptance criteria' in h or 'description' in h or 'task name' in h:
            desc_col_idx = desc_col_idx or c_idx # prioritize the first one found or we could check specific ones
        elif 'impact' in h:
            impact_col_idx = c_idx
        elif 'reason' in h:
            reason_col_idx = c_idx
        elif 'solution' in h:
            solution_col_idx = c_idx
            
    if rag_col_idx is None:
        continue
        
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row:
            continue
            
        rag_val = row[rag_col_idx]
        if rag_val in ['Red', 'Amber']:
            issues.append({
                'sheet': sheet_name,
                'criteria': str(row[desc_col_idx]) if desc_col_idx is not None else '',
                'rag': rag_val,
                'impact': str(row[impact_col_idx]) if impact_col_idx is not None else '',
                'reason': str(row[reason_col_idx]) if reason_col_idx is not None else '',
                'solution': str(row[solution_col_idx]) if solution_col_idx is not None else ''
            })

print(json.dumps(issues, indent=2))
