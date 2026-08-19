import openpyxl
import json

wb = openpyxl.load_workbook(r"C:\Users\User\Documents\Innovant\Internship\Project Dev\Development\AI Resume Analyser\testing report\InnovEdge_Plan_AI_Resume_Analyzer_RAG Report.xlsx")

sheets_to_read = ['Milestones', 'Deliverables', 'Detailed Plan', 'PM Reviews']

extracted_data = {}

for sheetname in sheets_to_read:
    if sheetname not in wb.sheetnames:
        print(f"Sheet {sheetname} not found!")
        continue
    sheet = wb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    
    # Let's find the column index for Acceptance Criteria (allowing for typos like Acceptacne Criteia)
    ac_idx = None
    for idx, h in enumerate(headers):
        if h and any(word in str(h).lower() for word in ['acceptance', 'acceptacne', 'criteria', 'criteia']):
            ac_idx = idx
            break
            
    if ac_idx is None:
        print(f"Could not find Acceptance Criteria column in {sheetname}")
        continue
        
    print(f"Sheet: {sheetname}, Acceptance Criteria column header: {headers[ac_idx]} (index {ac_idx})")
    
    extracted_data[sheetname] = []
    for r_idx, r in enumerate(rows[1:], start=2):
        # Only extract rows that have some content
        if any(r):
            row_dict = {}
            for i in range(min(len(headers), len(r))):
                if headers[i] is not None:
                    val = r[i]
                    if hasattr(val, 'isoformat'):
                        val = val.isoformat()
                    row_dict[str(headers[i])] = val
            extracted_data[sheetname].append({
                "row_index": r_idx,
                "row_data": row_dict
            })

with open("extracted_criteria.json", "w", encoding="utf-8") as f:
    json.dump(extracted_data, f, indent=2, ensure_ascii=False)

print("Done! Extracted data written to extracted_criteria.json")

